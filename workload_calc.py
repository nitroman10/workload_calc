import openpyxl as opxl
import iteround
import numpy as np


def get_min_max_col(sheet):
    start_col_num = 0
    end_col_num = 0
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == 'Должность':
                start_col_num = cell.column + 1
            if cell.value == 'Часы':
                end_col_num = cell.column - 2
    return start_col_num, end_col_num


def replace_row(table, row, row_num):
    table[row_num] = row
    return table


def replace_column(table, column, col_num):
    for i in len(table):
        table[i][col_num] = column[i]
    return table


def get_min_max_row(sheet):
    row_num_list = []
    for cell in sheet.iter_cols(min_col=4, max_col=4):
        row_num_list = [
            cell.row for cell in cell if cell.value == '08404 отдел']
    return row_num_list[0], row_num_list[-2]


def get_value_name_column(worksheet, name, column_n):
    return(worksheet.cell(names.index(name) + min_row, column_n).value)


def get_workfield():
    work_field = []
    for i, row in enumerate(sheet.iter_rows(min_row, max_row)):
        work_field.append([cell.value for cell in row]
                          [min_col - 1: max_col])
    return work_field


def get_workers_load(name, min_col, max_col):
    s = 0
    for column_n in range(min_col, max_col+1):
        value = get_value_name_column(name, column_n)
        if value:
            s += value
    return(s, tab_worktime_vector[names.index(name)])


def get_workers_overload_vect(name):
    worker_load, worker_tab_load = get_workers_load(name)
    workers_overload = worker_load - worker_tab_load
    themes_workers_overload_vect = []

    for column_n in range(min_col, max_col+1):
        value = get_value_name_column(name, column_n)
        themes_workers_overload_vect.append(
            (value/worker_load) * workers_overload if value else 0)

    return themes_workers_overload_vect


def get_theme_load_vector(theme):
    load_vector = []
    column = themes_cols[theme]
    for name in names:
        load_vector.append(get_value_name_column(name, column))
    return load_vector


def print_overall_load():
    s = 0
    for name in names:
        print(name, get_workers_load(name))
        s += get_workers_load(name)[0]
    print(s)


def set_aver_load_field(table, themes_load_vector, names):

    for j, load in enumerate(themes_load_vector):
        aver_load = []
        theme_vector = []
        if load:
            loaded_workers_qty = len(
                [np_vector[j] for np_vector in table if np_vector[j] != 0])
            aver_load = abs(load/loaded_workers_qty)
            for i, name in enumerate(names):
                theme_vector.append(aver_load if table[i][j] != 0 else 0.0)
        for i in range(len(theme_vector)):
            table[i][j] = theme_vector[i]
    return(table)


def write_main_workfield(table):
    for i, row in enumerate(main_worksheet.iter_rows(min_row=min_row,
                                                     max_row=max_row,
                                                     min_col=min_col,
                                                     max_col=max_col)):
        for j, cell in enumerate(row):
            cell.value = table[i][j]


def get_over_under_load_matrix(table):
    global names_qty, tab_worktime_vector
    overload_matrix = []
    for i in range(names_qty):
        name_load_vector = [num for num in table[i]]
        load_dif = sum(name_load_vector) - tab_worktime_vector[i]
        overload_vector = [load_dif*(num/sum(name_load_vector))
                           for num in name_load_vector]
        overload_matrix.append(overload_vector)
    return(overload_matrix)


def balance_workfield(table):
    global names_qty, themes_qty
    overunderload_matrix = get_over_under_load_matrix(table)
    for current_column in reversed(range(themes_qty - 1)):
        if themes_load_vector[current_column] == 0:
            continue
        over_underload_theme_vector = [vector[current_column]
                                       for vector in overunderload_matrix]
        # print(sum(over_underload_theme_vector))
        for current_row in range(names_qty):
            if over_underload_theme_vector[current_row] > 0:
                table[current_row][current_column] -= over_underload_theme_vector[current_row]
                underload_worker_qty = [
                    np_vector for np_vector in over_underload_theme_vector if np_vector < 0]
                s = abs(sum(underload_worker_qty))
                for worker_num in range(len(over_underload_theme_vector)):
                    if worker_num == current_row or over_underload_theme_vector[worker_num] > 0:
                        continue
                    else:
                        table[worker_num][current_column] += (abs(
                            over_underload_theme_vector[current_row]*(over_underload_theme_vector[worker_num]/s)) if s != 0 else 0)
    return(table)


def set_none_to_zero(table):
    for i, row in enumerate(table):
        for j, item in enumerate(row):
            if item is None or item == 0:
                table[i][j] = 0.0
    return table


def set_zero_to_none(table):
    for i, row in enumerate(table):
        for j, item in enumerate(row):
            if item == 0.0:
                table[i][j] = None
    return table


def round_field(table):
    for col_num in range(themes_qty):
        if themes_load_vector[col_num] != 0:
            theme_vector = [table[i][col_num] for i in range(names_qty)]
            theme_vector = iteround.saferound(
                theme_vector, 2, strategy='largest', topline=abs(themes_load_vector[col_num]))
        else:
            theme_vector = [0.0 for _ in range(names_qty)]
        for i in range(len(theme_vector)):
            table[i][col_num] = theme_vector[i]
    for i, row in enumerate(table):
        table[i] = iteround.saferound(
            row, 2, strategy='largest', topline=tab_worktime_vector[i])
    return table


def round_retain_sum(np_vector):
    np_vector = np_vector*100  # We want 2 decimal precision
    N = np.round(np.sum(np_vector)).astype(int)
    y = np_vector.astype(int)
    M = np.sum(y)
    K = N - M
    z = y-np_vector
    if K != 0:
        idx = np.argpartition(z, K)[:K]
        y[idx] += 1
    return y/100


def set_single_theme_load_vectors(table):
    single_theme_rows = []
    for i, row in enumerate(table):
        if len([num for num in row if num > 0]) == 1:
            single_theme_rows.append(i)
    for i in single_theme_rows:

        pass
    return table


def main():
    global themes_qty, names_qty, filepath, workbook, sheet, min_row, max_row, min_col, max_col, tab_worktime_vector, themes_load_vector, main_workbook, main_worksheet
    filepath = '084.xlsx'

    workbook = opxl.load_workbook(filepath, data_only=True)
    sheet = workbook.active
    min_row, max_row = get_min_max_row(sheet)
    min_col, max_col = get_min_max_col(sheet)

    main_workbook = opxl.load_workbook(filepath)
    main_worksheet = main_workbook.active
    themes = [sheet.cell(1, i).value for i in range(min_col, max_col+1)]
    themes_cols = dict([(sheet.cell(1, i).value, i)
                        for i in range(min_col, max_col+1)])
    names = [sheet.cell(i, 5).value for i in range(min_row, max_row+1)]
    names_rows = [(sheet.cell(i, 5).value, i)
                  for i in range(min_row, max_row+1)]
    themes_load_vector = [sheet.cell(
        min_row-1, i).value for i in range(min_col, max_col+1)]
    themes_load_vector = [
        num if num is not None else 0.0 for num in themes_load_vector]
    tab_worktime_vector = [sheet.cell(
        i, max_col+2).value for i in range(min_row, max_row+1)]
    active_dep_themes = [theme_num for theme_num in list(
        zip(themes, themes_load_vector)) if theme_num[1] != 0]
    themes_qty = len(themes)
    names_qty = len(names)

    workfield = np.array(get_workfield())
    workfield = set_aver_load_field(workfield, themes_load_vector, names)
    workfield = set_none_to_zero(workfield)

    workfield = set_single_theme_load_vectors(workfield)
    workfield = balance_workfield(workfield)
    # for _ in range(10):
    #     workfield = balance_workfield(workfield)
    # for _ in range(20):
    # workfield = round_field(workfield)
    print(*workfield[8:9], sep='\n\n', end='\n\n')
    # workfield = set_zero_to_none(workfield)
    write_main_workfield(workfield)

    main_workbook.save('new.xlsx')


if __name__ == '__main__':
    main()
